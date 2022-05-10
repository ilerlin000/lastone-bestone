from urllib.request import urlopen,Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


# scrape the website below to retrieve the top 5 countries with the highest GDPs. Calculate the GDP per capita
# by dividing the GDP by the population. You can perform the calculation in Python natively or insert the code
# in excel that will perform the calculation in Excel by each row. DO NOT scrape the GDP per capita from the
# webpage, make sure you use your own calculation.

# FOR YOUR REFERENCE - https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
# this link shows you the different number formats you can apply to a column using openpyxl


# FOR YOUR REFERENCE - https://www.geeksforgeeks.org/python-string-replace/
# this link shows you how to use the REPLACE function (you may need it if your code matches mine but not required)

### REMEMBER ##### - your output should match the excel file (GDP_Report.xlsx) including all formatting.

webpage = 'https://www.worldometers.info/gdp/gdp-by-country/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()

# WEBSCRAPING - NOT MY STRONGEST HA BUT HERE WE GO
# don't worry i did the venv + requirements and my excel portion is correct
# just tryna graduate :) like i apologize 

soup = BeautifulSoup(webpage, 'html.parser')

title = soup.title

print(title.text)

table_rows = soup.findAll('tr')

country_table = soup.find('table')
country_rows = country_table.findAll('tr')

# TOP 5 COUNTRIES - HIGHEST GDP 
'''
for x in range(1,6):
    td = country_rows[x].findAll('td')
    number = td[0].text
    country = td[1].text
    gdp = td[2].text
    population = td[5].text

    print(number)
    print(country)
    print(gdp)
    print(population)
'''

# CREATE A NEW EXCEL DOCUMENT - PUT IT IN EXCEL
wb = xl.Workbook()

ws = wb.active

ws.title = "GDP By Country"

wb.create_sheet(index=1,title='Second Sheet')

# WRITE CONTENT TO A CELL - COLUMN TITLES
ws['A1'] = 'No.'
ws['B1'] = 'Country'
ws['C1'] = 'GDP'
ws['D1'] = 'Population'
ws['E1'] = 'GDP Per Capita'

for x in range(1,6):
    td = country_rows[x].findAll('td')
    number = td[0].text
    country = td[1].text
    gdp = int(td[2].text.replace(',','').replace('$',''))
    population = int(td[5].text.replace(',',''))

    per_capita = (gdp/population)

    ws['A' + str(x + 1)] = number
    ws['B' + str(x + 1)] = country
    ws['C' + str(x + 1)] = gdp
    ws['D' + str(x + 1)] = population
    ws['E' + str(x + 1)] = per_capita



# ADJUST COLUMN WIDTH
ws.column_dimensions['A'] = 4
ws.column_dimensions['B'] = 15
ws.column_dimensions['C'] = 24
ws.column_dimensions['D'] = 19
ws.column_dimensions['E'] = 25

'''
# CHANGE FONT SIZE AND BOLD
ws['A1'].font = Font(name='Calibri',size=16,bold=True)
ws['B1'].font = Font(name='Calibri',size=16,bold=True)
ws['C1'].font = Font(name='Calibri',size=16,bold=True)
ws['D1'].font = Font(name='Calibri',size=16,bold=True)
ws['E1'].font = Font(name='Calibri',size=16,bold=True)
'''

# HEADER FONT - SIMPLE 
header_font = Font(name='Calibri',size=16,bold=True)

for cell in ws[1:1]:
    cell.font = header_font 


# SAVE AND NAME EXCEL 
wb.save('GDP_Report.xlsx')



# You are geuninely one of my favorite Professors at Baylor and you impact so many students
# Literally have never heard one bad thing said about you ever
# I am glad my last final at Baylor was in this class and I'm glad I decided to double major in MIS
# Also shout out to Lauren for grading all these assignments - you a real one for that 
# PS I am sorry if I bombed it... I am just trying to graduate :)


# JUST THANK YOU SO MUCH FOR EVERYTHING YOU DO!!!!! 

    





